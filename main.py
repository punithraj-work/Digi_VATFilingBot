#Program to open the web browser with URL and read data from webpage and dump into an excel "outputfile"

# from typing import Text
import json,os
from site import USER_BASE
# from numpy.f2py.auxfuncs import replace
# from numpy.f2py.crackfortran import true_intent_list
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

import time
from datetime import datetime
# from time import sleep 
# import openpyxl # Used to write to a new excel file
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# import calendar
from pythontimerdisplay import countdown
# from TaxciseAPI import gettaxcisedata,countdecl,funcdatarecon
from TaxciseAPI import gettaxcisedata,countdecl
from createexcel import createexcelfile 
from Gui import pyGUI
import pandas as pd
import numpy as np
import chromedriver_autoinstaller as chrome

from readpdf import readpdf

def getcurrenttime():
    timestamp = datetime.now()
    timestamp = timestamp.strftime("%d-%b %H:%M:%S")
    return timestamp

def checkelembyxpath(par1):

    try:
        #driver.find_element_by_id(par1)
        driver.find_element_by_xpath(par1)
    except NoSuchElementException:
        return False
    else:
        return True

def checkelembyid(par1):

    try:
        driver.find_element_by_id(par1)
    except NoSuchElementException:
        print("No element found")
        return False
    else:
        print("element found")
        return True

def takeelementval(x,y,z):
    driver.find_element(By.XPATH, "//"+str(x)+"[@"+str(y)+"="+str(z)+"]")


def formloadinterval(t):
    countdown(int(t))
    # startread = input("Start System to Read?:(press 'y' and press 'enter' to continue)")   #"readonly3@taxcise.ae"

def fillpdfdata():
    global rowcommonid,row2id,row3id,row4id,row5id,row6id,row7id,row9id,row10id,amountid,vatid,adjstid,form201s1,count

    rowcommonid = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblStandardRatedSuppliesIn"
    row2id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblTaxRefundsProvidedToTourists"
    row3id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblSuppliesSubjectToTheReverseChargeProvisions"
    row4id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblZeroRatedSupplies"
    row5id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblExemptSupplies"
    # row6id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblGoodsImportedIntoThe"
    row7id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblAdjustmentsAndAdditionsToGoodsImported"
    
    row9id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblStandardRatedExpenses"
    row9idb = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblStandardRatedExpensesRecoverable"
    row10id = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblSuppliesSubjectToTheReverseChargeProvisionsExpenses"
    row10idb = "p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_lblSuppliesSubjectToTheReverseChargeProvisionsExpensesRecoverable"
    
    amountid = "AmountValue"
    vatid = "VatAmountValue"
    adjstid = "AdjustmentValue"
    form201s1 = driver.find_elements(By.XPATH, "//div[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_pnlVatReturn' and @style='border-collapse:collapse;']/div")
    # for line in range(1, len(form201s1)):
    
    count=1
    while(count==1):
        #VAT on Sales and All Other Outputs Table
        # For clear(), check this link "https://stackoverflow.com/questions/70125902/how-to-clear-text-field-with-selenium-python"

        #For 1a
        a1amount = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"AbuDhabi"+str(amountid)+"']")
        a1amount.clear()
        a1amount.send_keys(pdfdata[9])
        a1vat = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"AbuDhabi"+str(vatid)+"']")
        a1vat.clear()
        a1vat.send_keys(pdfdata[10])
        a1adjs = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"AbuDhabi"+str(adjstid)+"']")
        a1adjs.clear()
        a1adjs.send_keys(pdfdata[11])
        
        #For 1b
        b1amount = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Dhabi"+str(amountid)+"']")
        b1amount.clear()
        b1amount.send_keys(pdfdata[14])
        b1vat = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Dhabi"+str(vatid)+"']")
        b1vat.clear()
        b1vat.send_keys(pdfdata[15])
        b1adjs = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Dhabi"+str(adjstid)+"']")
        b1adjs.clear()
        b1adjs.send_keys(pdfdata[16])

        #For 1c
        c1amount = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Sharjah"+str(amountid)+"']")
        c1amount.clear()
        c1amount.send_keys(pdfdata[19])
        c1vat = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Sharjah"+str(vatid)+"']")
        c1vat.clear()
        c1vat.send_keys(pdfdata[20])
        c1adjs = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Sharjah"+str(adjstid)+"']")
        c1adjs.clear()
        c1adjs.send_keys(pdfdata[21])

        #For 1d
        d1amount = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Ajman"+str(amountid)+"']")
        d1amount.clear()
        d1amount.send_keys(pdfdata[24])
        d1vat = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Ajman"+str(vatid)+"']")
        d1vat.clear()
        d1vat.send_keys(pdfdata[25])
        d1adjs = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Ajman"+str(adjstid)+"']")
        d1adjs.clear()
        d1adjs.send_keys(pdfdata[26])

        #For 1e
        e1amount = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"UmmAlQuwain"+str(amountid)+"']")
        e1amount.clear()
        e1amount.send_keys(pdfdata[29])
        e1vat = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"UmmAlQuwain"+str(vatid)+"']")
        e1vat.clear()
        e1vat.send_keys(pdfdata[30])
        e1adjs = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"UmmAlQuwain"+str(adjstid)+"']")
        e1adjs.clear()
        e1adjs.send_keys(pdfdata[31])

        #For 1f
        f1amount = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"RasAlKhaimah"+str(amountid)+"']")
        f1amount.clear()
        f1amount.send_keys(pdfdata[34])
        f1vat = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"RasAlKhaimah"+str(vatid)+"']")
        f1vat.clear()
        f1vat.send_keys(pdfdata[35])
        f1adjs = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"RasAlKhaimah"+str(adjstid)+"']")
        f1adjs.clear()
        f1adjs.send_keys(pdfdata[36])

        #For 1g
        g1amount = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Fujairah"+str(amountid)+"']")
        g1amount.clear()
        g1amount.send_keys(pdfdata[39])
        g1vat = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Fujairah"+str(vatid)+"']")
        g1vat.clear()
        g1vat.send_keys(pdfdata[40])
        g1adjs = driver.find_element(By.XPATH, "//input[@id='"+str(rowcommonid)+"Fujairah"+str(adjstid)+"']")
        g1adjs.clear()
        g1adjs.send_keys(pdfdata[41])

        #For 2
        # row2amount = driver.find_element(By.XPATH, "//input[@id='"+str(row2id)+str(amountid)+"']")
        # row2amount.clear()
        # row2amount.send_keys(pdfdata[45])
        # row2vat = driver.find_element(By.XPATH, "//input[@id='"+str(row2id)+str(vatid)+"']")
        # row2vat.clear()
        # row2vat.send_keys(pdfdata[46])
        # # row2adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row2id)+str(adjstid)+"']")
        # # row2adjs.send_keys(pdfdata[47])

        #For 3
        row3amount = driver.find_element(By.XPATH, "//input[@id='"+str(row3id)+str(amountid)+"']")
        row3amount.clear()
        row3amount.send_keys(pdfdata[50])
        row3vat = driver.find_element(By.XPATH, "//input[@id='"+str(row3id)+str(vatid)+"']")
        row3vat.clear()
        row3vat.send_keys(pdfdata[51])
        # row3adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row3id)+str(adjstid)+"']")
        # row3adjs.send_keys(pdfdata[52])

        #For 4
        row4amount = driver.find_element(By.XPATH, "//input[@id='"+str(row4id)+str(amountid)+"']")
        row4amount.clear()
        row4amount.send_keys(pdfdata[55])
        # row4vat = driver.find_element(By.XPATH, "//input[@id='"+str(row4id)+str(vatid)+"']")
        # row4vat.send_keys(pdfdata[9])
        # row4adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row4id)+str(adjstid)+"']")
        # row4adjs.send_keys(pdfdata[9])

        #For 5
        row5amount = driver.find_element(By.XPATH, "//input[@id='"+str(row5id)+str(amountid)+"']")
        row5amount.clear()
        row5amount.send_keys(pdfdata[58])
        # row5vat = driver.find_element(By.XPATH, "//input[@id='"+str(row5id)+str(vatid)+"']")
        # row5vat.send_keys(pdfdata[9])
        # row5adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row5id)+str(adjstid)+"']")
        # row5adjs.send_keys(pdfdata[9])

        #For 6
        # row6amount = driver.find_element(By.XPATH, "//input[@id='"+str(row6id)+str(amountid)+"']")
        # row6amount.send_keys(pdfdata[9])
        # row6vat = driver.find_element(By.XPATH, "//input[@id='"+str(row6id)+str(vatid)+"']")
        # row6vat.send_keys(pdfdata[9])
        # # row6adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row6id)+str(adjstid)+"']")
        # # row6adjs.send_keys(pdfdata[9])
    
        #For 7
        row7amount = driver.find_element(By.XPATH, "//input[@id='"+str(row7id)+str(amountid)+"']")
        row7amount.clear()
        row7amount.send_keys(pdfdata[65])
        row7vat = driver.find_element(By.XPATH, "//input[@id='"+str(row7id)+str(vatid)+"']")
        row7vat.clear()
        row7vat.send_keys(pdfdata[66])
        # row7adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row7id)+str(adjstid)+"']")
        # row7adjs.send_keys(pdfdata[9])

        #VAT on Expenses and All Other Inputs
        row9amount = driver.find_element(By.XPATH, "//input[@id='"+str(row9id)+str(amountid)+"']")
        row9amount.clear()
        row9amount.send_keys(pdfdata[88])
        row9vat = driver.find_element(By.XPATH, "//input[@id='"+str(row9idb)+str(vatid)+"']")
        row9vat.clear()
        row9vat.send_keys(pdfdata[89])
        row9adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row9id)+str(adjstid)+"']")
        row9adjs.clear()
        row9adjs.send_keys(pdfdata[90])
        
        row10amount = driver.find_element(By.XPATH, "//input[@id='"+str(row10id)+str(amountid)+"']")
        row10amount.clear()
        row10amount.send_keys(pdfdata[93])
        row10vat = driver.find_element(By.XPATH, "//input[@id='"+str(row10idb)+str(vatid)+"']")
        row10vat.clear()
        row10vat.send_keys(pdfdata[94])
        # row10adjs = driver.find_element(By.XPATH, "//input[@id='"+str(row10id)+str(adjstid)+"']")
        # row10adjs.send_keys(pdfdata[9])

        emailid = driver.find_element(By.XPATH, "//input[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_txtDeclerantEmailAddress']")
        emailid.clear()
        emailid.send_keys(usernameI)

        checkboxdo = driver.find_element(By.XPATH, "//input[@id='chkAccept']")
        checkboxdo.click()

        saveasdraft = driver.find_element(By.XPATH, "//input[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_Filing_btnSaveAsDraft']")
        saveasdraft.click()
        count+=1
        
        # if os.path.exists("test.txt"):
        #     os.remove("test.txt")

def writetoexcel(p1,p2):
    #Converting array into dataframe
    df = pd.DataFrame(p1)
    #Updating to an Excel
    # wb = load_workbook("outputfile.xlsx")
    wb = load_workbook("Declaration Number Reconcilation File.xlsx")
    # wb = load_workbook("D:\Punith\Python\Python WebScrapping\Declarations Count\Declaration Number Reconcilation File.xlsx")
    # Select First Worksheet
    ws = wb.worksheets[p2]
    #ws = wb['Sheet1']
    for items in dataframe_to_rows(df, index=False, header=False):
        ws.append(items)
    wb.save("Declaration Number Reconcilation File.xlsx")
    # wb.save("D:\Punith\Python\Python WebScrapping\Declarations Count\Declaration Number Reconcilation File.xlsx")
    wb.close()

def openform(openformI):
    transactionc=0
    if openformI=="y":
        transactionc = (driver.find_element_by_xpath("//*[@id='showTotalRecords']/span[1]").text)
    else:
        pass
    return transactionc

def chkblockui():
    # if checkelembyxpath("div[@class='blockUI blockMsg blockPage']"):
    if checkelembyxpath("//div[@class='blockUI']"):
        countdown(10)
        print("please wait, the page is loading...")
    else:
        pass
    if checkelembyxpath("//div[@class='blockUI']"):
        chkblockui()


sysstarttime = datetime.now()
sysstarttime = sysstarttime.strftime("%H:%M:%S")
# print(sysstart)

print("   HI!!! Welcome to Recon BOT\n *********************************\n BOT Started at %s \n" % sysstarttime)

def main():
    global taxcisejson,driver,taxcisefname,sysendtime,usernameI,pdfdata
    if os.path.exists("./PDF/example.pdf"):
        pdfdata = readpdf()
    else:
        print("PDF File Not Found, BOT is terminated")
    userprompt = "2" #GUIaccept[0]   #input("Do you want to run Automatic or Manual Process:?\n 1=Automatic\n 2=Manual\n ")
    formname = "1" #GUIaccept[1]   #input("Select the form name:?\n 1=EX201 - Excise Goods that require Customs clearance\n 2=EX202A – Designated Zone Reporting\n 3=EX202B – Producer Declaration\n 4=Inventory - EX203A - Local Purchase Form\n 5=Inventory - EX203B - Lost and Damaged Declaration\n 6=Inventory – EX203C – Transfer of Ownership Within Designated Zones\n 7=EX203 - Deductible Excise Tax\n")

    if formname=="1" or formname=="2" or formname=="3" or formname=="5" or formname=="7":
        formperiod= "10-2022" #GUIaccept[2]   #input("Please Enter the Period in the format M-YYYY / MM-YYYY (Eg: 1-2021 / 10-2021 :\n)")
    elif formname=="6":
        print("By Default BOT will filter the Approved Status for FORM 203C")
    else:
        print("Forms will no periods is selected")


    if userprompt=="1":
        print("BOT run for FTA")
        print("Automatic Process enabled.. Open the Excise Dashboard")
        usernameI = "prashant@digiwabbit.com"
        passwordI = "Hello007^"
    elif userprompt=="2":
        print("BOT run for FTA")
        print("Manual Process enabled.. Open the Excise Dashboard")
        usernameI = "prashant@digiwabbit.com"
        passwordI = "Hello007^"
    else:
        print("invalid Seletion..!! BOT is terminated")
        exit()

    global clientname,TStatus,TStatusF,projectname,userdateselect,userdate
    
    # with open('link', 'r')as f:
    #     linkdata = json.load(f)

    clientname="1" #linkdata['client']
    projectname="1" #linkdata['link']
    # clientname=GUIaccept[3]   #input("You are running for which user?\n 1 Alfakhar\n 2 Steinweg\n")
    Generatekey="n" #GUIaccept[4]   #input("Do you want to download data from taxcise? (y/n) : ")
    TStatus="1" #GUIaccept[6]

    # if TStatus=="1":
    #     TStatusF="Approved"
    # else:
    #     TStatusF="Drafted"

    if Generatekey=='y':
        taxcisejson=gettaxcisedata(clientname,0,projectname)

      
    # chrome.get_chrome_version()
    chromepath=chrome.install(True) # Check if the current version of chromedriver exists
                                        # and if it doesn't exist, download it automatically,
                                        # then add chromedriver to path
                                        # If you pass 'TRUE' parameter then it will download the 
                                        # Chrome driver in Current Working Path
                                        

    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Chrome(options=options ,executable_path=chromepath)
    driver.get("https://eservices.tax.gov.ae/en-us/")

    username = driver.find_element_by_xpath("//input[@placeholder = 'Email address']")
    password = driver.find_element_by_xpath("//input[@placeholder = 'Password']")


    username.send_keys(usernameI)
    password.send_keys(passwordI)
    #scodeinp.send_keys(str(securitycode.text))

    countdown(int(25))
    # startread = input("Start System to Read?:(press 'y' and press 'enter' to continue)")   #"readonly3@taxcise.ae"

    '''
    #Code to automatically click on taxable person
    accesstaxable = driver.find_element_by_xpath("//a[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_MOFTaxableUserDashboard_gv_TaxablePersons_ctl02_btnLoginTaxablePerson']")
    accesstaxable.click()
    alert = driver.switch_to_alert()
    alert.accept()
    '''
    
    chkblockui()
    if checkelembyxpath("//a[@href='/User/Vatreturn/dashboard']"):
        countdown(10) 
    excisedashb = driver.find_element_by_xpath("//a[@href='/User/Vatreturn/dashboard']")
    excisedashb.click()
    chkblockui()
    Currenturl = driver.current_url
    taxcisefname = ""
    if userprompt=="2":
        # tabs = driver.find_elements(By.XPATH, "//ul[@class='nav nav-pills']/li")
        # print(len(tabs))
        if formname=="1":
            # form201clk = driver.find_element(By.XPATH, "//ul[@class='nav nav-pills']/li[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_MOFExciseDeclationDashboard_ucDashboardTabs_liExciseImport']/a[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_MOFExciseDeclationDashboard_ucDashboardTabs_tabExciseImport']")
            form201clk = driver.find_element(By.XPATH, "//a[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_VATReturnDashboard_tabVATReturn']")
            form201clk.click()
            if checkelembyid('p_lt_ctl10_pageplaceholder_p_lt_ctl00_VATReturnPage_btnVatTaxReturn'):
                newvatfile = driver.find_element_by_id("p_lt_ctl10_pageplaceholder_p_lt_ctl00_VATReturnPage_btnVatTaxReturn")
                newvatfile.click()
            else:
                editvatfile = driver.find_element(By.XPATH, "//table[@id='p_lt_ctl10_pageplaceholder_p_lt_ctl00_VATReturnPage_gv_VATTaxReturn']/tbody/tr[2]/td[11]/a")
                editvatfile.click()
            time.sleep(1)
            chkblockui()
            fillpdfdata()
        else:
            print("Invalid Selection\n")
            pass


    sysendtime = datetime.now()
    sysendtime = sysendtime.strftime("%H:%M:%S")
    # print("BOT ended at %s" % sysendtime)
    countdown(2)
    logout = driver.find_element(By.XPATH, "//a[@title='Logout']")
    logout.click()

    # time.sleep(15)
    countdown(int(5))
    # print("Please close the browser once logged out successfully")
    driver.close()

    '''
    #Converting array into dataframe
    df = pd.DataFrame(tdetails1)
    #Updating to an Excel
    wb = load_workbook("Declaration Number Reconcilation File.xlsx")
    # Select First Worksheet
    ws = wb.worksheets[0]
    #ws = wb['Sheet1']
    for items in dataframe_to_rows(df, index=False, header=False):
        ws.append(items)
    wb.save("Declaration Number Reconcilation File.xlsx")
    wb.close()
    '''


    #Code to open web browser
    '''
    import webbrowser
    webbrowser.open('https://eservices.tax.gov.ae/en-us/taxable/excise-dashboard/excise-goods-import-dashboard?TaxpayerUserGuid=ca93f69e-8482-423f-a2d8-da3e6e03df1e', new = 2)

    #Convert nos to string
    value = '2,376,000.00'
    # Convert number string with comma to integer object
    num = float(value.replace(',',''))
    num = num+num
    num = int(num)
    print(num)
    '''


    '''
    code to remove unique elements in two list
    x = [1, 2, 3, 4]

    f = [1, 11, 22, 33, 44, 3, 4]

    res = list(set(x) ^ set(f))

    print(sorted(res))

    [33, 2, 22, 11, 44]

    '''


# print(pyGUI())
import os
if os.path.exists("link"):
    if __name__=="__main__":
        main()
        # GUIaccept = pyGUI()
        # if GUIaccept==False:
        #     exit()
        # elif GUIaccept[5]=='Proceed':
        #     main(GUIaccept)
        #     # print('procceed')
else:
    if __name__=="__main__":
        main()
    print("Link file does not exist")


sysendtime = datetime.now()
sysendtime = sysendtime.strftime("%H:%M:%S")
print("BOT ended at %s" % sysendtime)