from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active
# ws.merge_cells('B2:F4')
# top_left_cell = ws['B2']
# thin = Side(border_style="thin", color="000000")
# double = Side(border_style="double", color="ff0000")
# top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
# top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
# top_left_cell.fill = fill = GradientFill(stop=("000000", "FFFFFF"))
# top_left_cell.font  = Font(b=True, color="FF0000")
# top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
# wb.save("styled.xlsx")


def createexcelfile():
    wb = Workbook()
    ws = wb.active
    ws.title="Declnos Count"
    cellrange = ws['A1:I1']
    cellno=0
    headings=['FTA FORM Info','Total Declarations (as per FTA)','Amount (as per FTA)','Period','Total Declarations (as per Taxcise)','Amount(as per Taxcise)','Diff in Decl nos','Diff in Value','Time Stamp (DD-MM HH:MM:SS)']
    for i in cellrange[0]:
        i.fill = PatternFill("solid", fgColor="00FFFF00")
        i.font  = Font(b=True, color="000000")
        i.value=headings[cellno]
        cellno+=1
        # i.value="Testing"


    ws1 = wb.create_sheet("recon BOT")
    ws1 = wb.worksheets[1]
    cellrange = ws1['A1:E1']
    cellno=0
    headings=['FTA FORM Info','Period','Decl no (missing on Taxcise)','Amount (as per FTA)','Time Stamp (DD-MM HH:MM:SS)']
    for i in cellrange[0]:
        i.fill = PatternFill("solid", fgColor="00FFFF00")
        i.font  = Font(b=True, color="000000")
        i.value=headings[cellno]
        cellno+=1

    wb.save("Declaration Number Reconcilation File.xlsx")
